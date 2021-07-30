"""Spreadsheet manipulation with openpyxl.

Demostration by Sadie Bartholomew
Held on 30/07/21 for the NCAS Python Technical Group meeting.

"""

# 1. Accessing openpyxl
# First make sure it is installed (`pip install openpyxl` or similar)!
import openpyxl

# 2. Load in your spreadsheet for manipulation with openpyxl.
#    Note "workbook" == spreadsheet in Excel speak! Hence the method name, etc.
workbook = openpyxl.load_workbook('demo-spreadsheet.xlsx')  # path argument

# 3. (Alternatively, you can create a whole new spreadsheet without having to
#     create it in the filesystem first, like so, but we don't do that here.)
new_workbook = openpyxl.Workbook()  # just to illustrate, we don't use this

# 4. Determine all of the tabs in the spreadsheet via iteration.
#    Note "worksheet" == tab of a spreadsheet in Excel speak!
#    (New tabs can also be created, but we don't try that here.)
for tab in workbook:
    print(tab.title)
# Alternatively there is a specialised method for this:
print(workbook.sheetnames)

# 5. Access a tab of interest by indexing, which we call the "active" tab
worksheet = workbook["Playground 1 (basic worksheet)"]

# 6. Access a given cell to read using indexing
print("Value is:", worksheet["B2"].value)  # read the data
# And read the style too!
print("Font properties are:", worksheet["B2"].font)
print("Background properties are:", worksheet["B2"].fill)

# 7. Accessing multiple cells is just as easy via indexing and slicing, e.g:
worksheet["B2:B6"]
for row in worksheet["B2:B6"]:  # all on column B, so down a column
    cell = row[0]  # take first and only cell in this column list
    print(f"{cell.value} is written here in colour {cell.font.color.rgb}")

# 8. Let's now try writing out data to the given tab
worksheet["C3"] = "Foo again"
# Important! It won't write the data out in-place, you need to save the
# workbook for the changes to be made... so let's save and take a look...

# 9. Try saving our amended workbook i.e. spreadsheet
workbook.save("demo-spreadsheet-edited.xlsx")

# 10. This time, let's try writing it out with a given style.
# First grab the style of the original Foo cell.
# This works to do it all in one line, even though it looks a bit clunky with
# the internal method and requires a copy so as not to affect the original cell
from copy import copy
blue_style_from_foo = copy(worksheet["B2"]._style)
# Then apply it to the newly-written Foo:
worksheet["C3"]._style = blue_style_from_foo
# Remember to save the changes for them to be applied!
workbook.save("demo-spreadsheet-edited.xlsx")

# 11. Let's write in one of the other styles, too, all in one go, and for
#     multiple cells too to mix it up:
green_style_from_bar = copy(worksheet["B4"]._style)
for row in worksheet["B8":"C20"]:  # iterating over both rows and columns now
    for cell in row:
        worksheet[cell.coordinate] = (
            f"This cell has coordinate {cell.coordinate}")
        worksheet[cell.coordinate]._style = green_style_from_bar
workbook.save("demo-spreadsheet-edited.xlsx")

# 12. Now to do something fun and pointless: let's make some spreadsheet art!

# a) Use an imaging library (Pillow) with NumPy to get a basic image array
from PIL import Image
import numpy as np

# b) Read in the PNG image to Pillow
# Original clip art image (used as an example) downloaded from:
#   http://clipart-library.com/clipart/leaf-clip-art-1.htm
image = Image.open("img/leaf-clip-art.png")
print(image.format, image.size, image.mode)  # confirms RGBA colours

# c) Convert the read-in image to an array
image_array = np.asarray(image)
print(image_array, image_array.size)

# d) It is a bit too big to be practical to view, so make it a bit smaller
image = image.reduce(3)
image_array = np.asarray(image)
print(image_array, image_array.size)
print(image_array[0,0], image_array.shape)

# e) Point to a different tab of the spreadsheet where the width and height
#    of the cells are equal, to create realistic pixels
art_worksheet = workbook["Playground 2 (square cells)"]

# f) Import matplotlib which helps with the pixel colour format conversions
import matplotlib.colors

# g) Now the main part: let's write the image as pixels on the spreadsheet tab!
#    Do this by iterating over the rows and columns respectively:
for row in range(image_array.shape[0]):
    for col in range(image_array.shape[1]):
        # i) Get the correct pixel from the array for this location
        pixel = image_array[row, col]

        # ii) Do colour conversions to get the colour format used by openpyxl
        rgb_tuple = pixel[:3]
        hexa = matplotlib.colors.rgb2hex([1.0*x/255 for x in rgb_tuple])
        colour = openpyxl.styles.colors.Color(rgb=f"FF{hexa.lstrip('#')}")
        print(hexa, colour.rgb)

        # iii) The main part: write corresponding array value as fill colour
        #      to form an effective image pixel on the spreadsheet.
        # Note the alternative syntax for accessing a row and column, which
        # uses numbers not letters for the column naming.
        art_worksheet.cell(
            row=row + 2, column=col + 2).fill = openpyxl.styles.PatternFill(
            fill_type='solid', fgColor=colour)  # '+ 2' to pad image on tab

# h) Finally save again to produce the pixel art on the filesystem spreadsheet
workbook.save("demo-spreadsheet-edited.xlsx")
